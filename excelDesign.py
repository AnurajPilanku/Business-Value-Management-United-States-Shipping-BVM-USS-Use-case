'''
Author  : Anuraj Pilanku

Usecase : BVM USS Report

'''
import string
import openpyxl
import sys
from openpyxl.styles import Alignment, Border,Side,Font,PatternFill
from copy import copy
from colorama import *

#inputs and outputs
path=sys.argv[1]
input=path+"\\"+"BVM_Report.xlsx"

#creating workbook and sheets
#workbook for laia
wb=openpyxl.load_workbook(input)
sh=wb.worksheets[0]
rowcount=sh.max_row
columncount=sh.max_column

#Development
darkGrey="A9A9A9"
silver="C0C0C0"
dimGrey="696969"
slateGrey="778899"






for i in range(1, columncount + 1):
    sh.cell(column=i,row=1).fill = PatternFill(start_color=dimGrey, end_color=dimGrey, fill_type="lightUp")

#setting cell borders
black = "000000"
thin = Side(border_style="thin", color=black)
#for i in range(1, columncount + 1):
for j in range(2, rowcount + 1):
            sh.cell(column=i,row=1).border = Border(top=thin, left=thin, right=thin,bottom=thin)
sh["AJ2"].value="=Median(AI2:AI"+str(rowcount)+")"
wb.save(path+"\\"+"BVMReport.xlsx")
print("success")





#python bvm.py C:\Users\2040664\anuraj\ipi2k\test.xlsx
#python lcs.py \\acdev01\3M_CAC\LCS_Borker\query_output\lcs.xlsx
