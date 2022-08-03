import openpyxl
import sys
import time

consolidatedpath=sys.argv[1]+"\\"+"BVM_Report.xlsx"
wb=openpyxl.load_workbook(consolidatedpath)
s=wb.active
#J-Allocation_Received-H
#M-Shipment_Date-M
#K-Planning_Releasing-I
#REQ-Ready to Ship CycleTime-AI
s["AI1"].value='Ready to Ship CycleTime'
s["AJ1"].value='KPI Value'
for row in range(2,s.max_row+1):
    if s["J"+str(row)].value in [None," "," "]:
        s["AI" + str(row)].value= "=M{rownum}-K{rownum}".format(rownum=row)#s["M"+str(row)].value-s["K"+str(row)].value
    else:
        s["AI" + str(row)].value ="=M{rownum}-J{rownum}".format(rownum=row) #s["M" + str(row)].value - s["J" + str(row)].value
#if allocation recieved=none, then {shipmentdate-planning releasing} , else {shipmentdate-Allocation recieved}#IF(J{rowNum}="",(M{rowNum}-K{rowNum}),(M{rowNum}-J{rowNum}))
wb.save(consolidatedpath)
time.sleep(5)

from win32com.client import Dispatch
import sys
import openpyxl
wkbk1 = sys.argv[1]

wb=openpyxl.load_workbook(consolidatedpath)
ws=wb.active
rowcount=ws.max_row
wb.close()

try:
    excel = Dispatch("Excel.Application")
    excel.Visible = 0
    source = excel.Workbooks.Open(consolidatedpath)
    # sh=source.worksheets(0)
    # used=sh.UsedRange
    excel.Range("AI2:AI"+str(rowcount)).Select()
    excel.Selection.Copy()
    # copy = excel.Workbooks.Open(wkbk2)
    excel.Range("AI2:AI"+str(rowcount)).Select()
    excel.Selection.PasteSpecial(Paste=-4163)
    source.Save()  # SaveAs(Filename:=sys.argv[1])
    source.Close()
    excel.Quit()
    print("success")
except:
    import os
    import win32com.shell.shell as shell
    import time

    # commands='taskkill /f /im EXCEL.EXE'
    # shell.ShellExecuteEx(lpVerb='runas',lpFile='cmd.exe',lpParameters='/c'+commands)
    os.system('taskkill /f /im EXCEL.EXE')
    time.sleep(10)

    excel = Dispatch("Excel.Application")
    excel.Visible = 0
    source = excel.Workbooks.Open(consolidatedpath)
    # sh=source.worksheets(0)
    # used=sh.UsedRange
    excel.Range("AI2:AI"+str(rowcount)).Select()
    excel.Selection.Copy()
    # copy = excel.Workbooks.Open(wkbk2)
    excel.Range("AI2:AI"+str(rowcount)).Select()
    excel.Selection.PasteSpecial(Paste=-4163)
    source.Save()  # SaveAs(Filename:=sys.argv[1])
    source.Close()
    excel.Quit()
    print("success")


