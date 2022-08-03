import pandas as  pd
import openpyxl
from xlrd.xldate import xldate_as_tuple
from datetime import datetime
import sys

input=sys.argv[1]
bvmwb=openpyxl.load_workbook(input)
bsh=bvmwb.worksheets[0]#['Mainsheet']
#removing #NA
for c in range(1,bsh.max_column+1):
    for r in range(2,bsh.max_row+1):
        if bsh.cell(column=c,row=r).value =="#NA":
            bsh.cell(column=c, row=r).value=bsh.cell(column=c,row=r).value.replace("#NA","")
#Removing other data in KPI Column
for r in range(3,bsh.max_row+1):
    bsh['Q'+str(r)].value=""
#Converting serial date to normal date format
reqCol='C:E:F:G:H:I:J:K:L:M:AF:AG'.split(":")
for collet in range(0,len(reqCol)):
    for r in range(2,bsh.max_row+1):
        val=bsh[reqCol[collet]+str(r)].value
        s = pd.Series([val])#([41516.43])
        dat=s.apply(lambda x: datetime(*xldate_as_tuple(x, 0)))
        reqdat=str(dat[0])
        bsh[reqCol[collet] + str(r)].value=reqdat
bvmwb.save(input)
bvmwb.close()
