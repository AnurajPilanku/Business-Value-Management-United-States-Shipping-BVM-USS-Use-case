#Anuraj Pilanku
#Apply formulae  from exceldata

import openpyxl
import os
import sys


mainpath=sys.argv[1]#r"C:\Users\2040664\anuraj\bvm\bvm_main_excel.xlsx"
formulaepath=sys.argv[2]#r"C:\Users\2040664\anuraj\bvm\test\formulaeDetailsaranged.xlsx"
mainsheet='Mainsheet'
mainwb=openpyxl.load_workbook(mainpath)
sheet=mainwb.worksheets[0]#mainwb[mainsheet]
form=openpyxl.load_workbook(formulaepath)
formsh=form.worksheets[0]

formulae_dict=dict()
for i in range(2,formsh.max_row-1):
    if formsh['E'+str(i)].value not in [None]:
        formulae_dict[formsh["A"+str(i)].value]="="+formsh['E'+str(i)].value.strip()
#print(formulae_dict)
for i in range(0,len(formulae_dict)):
    for j in range(2,sheet.max_row+1):
        #if j<10:
        sheet[list(formulae_dict.keys())[i]+str(j)].value=formulae_dict[list(formulae_dict.keys())[i]].format(rowNum=j)

        #print(list(formulae_dict.keys())[i]+str(j),formulae_dict[list(formulae_dict.keys())[i]].format(rowNum=j))
mainwb.save(mainpath)
print("success")
#python  C:\Users\2040664\PycharmProjects\pythonProject1\trimexcel.py C:\Users\2040664\anuraj\bvm\bvm_main_excel.xlsx C:\Users\2040664\anuraj\bvm\test\formulaeDetailsaranged.xlsx
