import openpyxl
import os
import sys
mainpath=sys.argv[1]#r"C:\Users\2040664\anuraj\bvm\test\bvm_main_excel.xlsx"
queryoutputpath=sys.argv[2]#r"C:\Users\2040664\anuraj\bvm\test\queryoutput"
queryonepath=sys.argv[3]#r"C:\Users\2040664\anuraj\bvm\test\queryoutput\query1.xlsx"
mainsheet='Mainsheet'
mainwb=openpyxl.load_workbook(mainpath)
sheet=mainwb[mainsheet]

#query0ne
queryonewb=openpyxl.load_workbook(queryonepath)
qonesheet=queryonewb.worksheets[0]
repeat=list()
for col in range(1,4):
    for row in range(2, qonesheet.max_row + 1):

        if qonesheet.cell(column=1,row=row).value not in repeat:
            if type(qonesheet.cell(column=col, row=row).value) in [str]:
                sheet.cell(column=col, row=row).value = qonesheet.cell(column=col, row=row).value.strip()
            else:
                sheet.cell(column=col, row=row).value = qonesheet.cell(column=col, row=row).value

        repeat.append(qonesheet.cell(column=1, row=row).value)
for row in range(2, qonesheet.max_row + 1):
    sheet.cell(column=2, row=row).value="=VLOOKUP(A{vl},query1!A:C,2,FALSE)".format(vl=row)
    sheet.cell(column=3, row=row).value="=VLOOKUP(A{vl},query1!A:C,3,FALSE)".format(vl=row)
queryonewb.close()

queryWbNames=os.listdir(queryoutputpath)
#print(queryWbNames)
for i in queryWbNames:
    if "~$" not in i:
        newsheet=mainwb.create_sheet(i.replace(".xlsx",""))
        queryWb=openpyxl.load_workbook(queryoutputpath+"\\"+i)
        qsheet=queryWb.worksheets[0]
        for col in range(1,qsheet.max_column+1):
            for row in range(1, qsheet.max_row + 1):
                if type(qsheet.cell(column=col,row=row).value) in [str]:
                    newsheet.cell(column=col,row=row).value=qsheet.cell(column=col,row=row).value.strip()
                else:
                    newsheet.cell(column=col, row=row).value = qsheet.cell(column=col, row=row).value
mainwb.save(mainpath)
print('success')

#python server.py C:\Users\2040664\anuraj\bvm\test\bvm_main_excel.xlsx C:\Users\2040664\anuraj\bvm\test\queryoutput