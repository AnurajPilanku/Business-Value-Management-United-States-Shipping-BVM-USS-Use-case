#python vlookup #ASN ODERNUM and Inventery
import openpyxl
import sys
import numpy

w=openpyxl.load_workbook(sys.argv[1])
s=w.active

try:

    def transfer(wbname, sourcekeyCol, sourceValCol, destcol):
        queryeight = openpyxl.load_workbook(sys.argv[2] + "\\" + wbname + ".xlsx")
        qet = queryeight.active
        eight = dict()
        for i in range(2, qet.max_row + 1):
            eight[str(qet[sourcekeyCol + str(i)].value).strip()] = qet[sourceValCol + str(i)].value
        for i in range(2, s.max_row + 1):
            if s["A" + str(i)].value not in [None, ""]:
                if str(s["A" + str(i)].value).strip() in list(eight.keys()):
                    s[destcol + str(i)].value = eight[str(s["A" + str(i)].value).strip()]

    transfer("query1", "A", "B", "B")
    transfer("query1", "A", "C", "C")

    #ASN ORDERS--colN
    aswb=openpyxl.load_workbook(sys.argv[2]+ "\\" + "query10" + ".xlsx")
    assh=aswb.worksheets[0]
    ordnumlist=[]
    for i in range(2,assh.max_row+1):
        if assh["A"+str(i)].value not in [None,""," "]:
            ordnumlist.append(assh["A"+str(i)].value.strip())
    for i in range(2,s.max_row+1):
        if s["A"+str(i)].value not in [None,""," "]:
            if s["A" + str(i)].value.strip() in ordnumlist:
                s["N" + str(i)].value = "Y"
            else:
                s["N" + str(i)].value = "N"
       
    #Inventry Col-O
    constantdatapath=sys.argv[3]
    invwb=openpyxl.load_workbook(constantdatapath+"\\"+"Inventory"+".xlsx")
    invsh=invwb.worksheets[0]
    ordnuminventrylist=[]
    for i in range(2,invsh.max_row+1):
        if invsh["A"+str(i)].value not in [None," ",""]:
            ordnuminventrylist.append(invsh["A"+str(i)].value.strip())
    for i in range(2,s.max_row+1):
        if s["B"+str(i)].value not in [None,""," "]:
            if s["B" + str(i)].value.strip() in ordnuminventrylist:
                s["O" + str(i)].value = "Y"
            else:
                s["O" + str(i)].value = "N"
       
    #KPI Value
    KPIQ=[]
    for i in range(2,s.max_row+1):
        if s["P"+str(i)].value not in [None,""," "]:
            KPIQ.append(s["P"+str(i)].value)
    s["Q2"].value=numpy.median(KPIQ)
    w.save(sys.argv[1])
    print("success")
except Exception as e:
    print(e)


#python server.py C:\Users\2040664\anuraj\bvm\test\queryoutput\query1.xlsx C:\Users\2040664\anuraj\bvm\test\queryoutput C:\Users\2040664\anuraj\bvm\test\bijoy.xlsx


